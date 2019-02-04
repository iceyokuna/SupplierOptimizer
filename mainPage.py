import sys
import time
from openpyxl import load_workbook
from PySide2.QtWebEngineWidgets import QWebEnginePage, QWebEngineView
from PySide2.QtUiTools import QUiLoader
from PySide2.QtWidgets import *
from PySide2.QtCore import *
from PySide2.QtGui import *
from SupplierController import *
from MapController import *

#Blueprint of main page
class MainPage(QObject):
    def __init__(self, parent=None):
        super(MainPage,self).__init__(parent)
        #Read File Ui
        ui_file = QFile('mainPage.ui')
        ui_file.open(QFile.ReadOnly)
        loader = QUiLoader()
        self.ui = loader.load('UI\mainPage.ui')
        self.window = self.ui.findChild(QWidget, 'mainWidget')
        palette = QPalette()
        palette.setBrush(QPalette.Background, QBrush(QPixmap("UI/background.jpg")))
        self.ui.setPalette(palette)
        
        ui_file.close()    

        #Setup mainpage 
        self.ui.setWindowTitle("LE Transport")

        #MainPage Attribute
        self.SupplierController = SupplierController()
        self.MapController = MapController()

        #Add Object
        ##Object List
        self.supplier_list = self.ui.findChild(QListWidget, 'supplierList')
        self.item_list = self.ui.findChild(QListWidget, 'itemList')
        self.supplier_list.clicked.connect(self.supplierListClicked)
        self.item_list.clicked.connect(self.itemListClicked)

        ##Object Button
        self.import_excel_button = self.ui.findChild(QPushButton, 'importExcelButton')
        self.start_button = self.ui.findChild(QPushButton, 'startButton')
        self.import_excel_button.clicked.connect(self.importExcel)
        self.start_button.clicked.connect(self.startCalculate)

        ##Object Text Edit
        self.customer_location = self.ui.findChild(QLineEdit, 'customer_location_edit')
        self.customer_location.setText("13.822515, 100.513089")

        ##Object Label
        self.supplier_label = self.ui.findChild(QLabel, 'supplier_label')
        self.supply_label = self.ui.findChild(QLabel, 'supply_label')
        self.transport_cost_label = self.ui.findChild(QLabel, 'transport_cost_label')
        self.supply_cost_label = self.ui.findChild(QLabel, 'supply_cost_label')
        self.time_label = self.ui.findChild(QLabel, 'time_label')
        self.distance_label = self.ui.findChild(QLabel, 'distance_label')

        ##Object Label2
        self.supplier_label_2 = self.ui.findChild(QLabel, 'supplier_label_2')
        self.supply_cost_label_2 = self.ui.findChild(QLabel, 'transport_cost_label_2')
        self.transport_cost_label_2 = self.ui.findChild(QLabel, 'supply_cost_label_2')
        self.distance_label_2 = self.ui.findChild(QLabel, 'distance_label_2')

        ##Object Label3
        self.supplier_label_3 = self.ui.findChild(QLabel, 'supplier_label_3')
        self.supply_cost_label_3 = self.ui.findChild(QLabel, 'transport_cost_label_3')
        self.transport_cost_label_3 = self.ui.findChild(QLabel, 'supply_cost_label_3')
        self.distance_label_3 = self.ui.findChild(QLabel, 'distance_label_3')
        
        #ตัวแปล ข้อมูลตัวอักษรเริ่มค้น
        self.supplier_str = "ชื่อซัพพลายเออร์ :"
        self.supply_str = "PRODUCT  "
        self.transport_cost_str = "ต้นทุนรวม  :"
        self.supply_cost_str = "ต้นทุนสินค้า   :"
        self.time_str = "เวลาในการคำนวน :"
        self.distance_str = "ระยะทาง  :"
        
        ##initialize web engine
        self.web_widget = self.ui.findChild(QWidget, 'map_widget')
        self.webEngineView = QWebEngineView()
        
        lat , lon = self.customer_location.text().split(',')
        self.MapController.setCustomerMarker(lat , lon)
        self.MapController.setSupplierMarker(lat , lon)
        mapHTML = self.MapController.getHTML()
        self.webEngineView.setHtml(mapHTML)
        self.webEngineView.resize(761, 721);
        self.webEngineView.setParent(self.web_widget)
        

    def importExcel(self):
        dialog = QFileDialog()
        filename = QFileDialog.getOpenFileName(dialog, "Import Supplier (.xlsx)",None, "Excel files (*.xlsx)")
        excel_location = filename[0]
        
        excel = load_workbook(excel_location)
        sheet = excel.get_sheet_by_name("Sheet1")
        sheet2 = excel.get_sheet_by_name("Sheet2")

        row = 2
        while(True):
            #If data is None then break the loop
            if(sheet.cell(row, 1).value is None):
                break
            self.SupplierController.addSupplier((sheet.cell(row, 1).value,
                                             sheet.cell(row, 2).value,
                                             sheet.cell(row, 3).value,
                                             sheet.cell(row, 4).value))
            row += 1

        row = 2
        distance_cost_list = []
        while(True):
            #If data is None then break the loop
            if(sheet2.cell(row, 1).value is None):
                break
            distance_cost_list.append((sheet2.cell(row, 1).value,
                                             sheet2.cell(row, 2).value))
            row += 1

        #print distance cost list
        for i in distance_cost_list:
            print(i)

        self.SupplierController.setDistanceCost(distance_cost_list)
        self.updateData()


    def updateData(self):
        self.supplier_list.clear()
        self.item_list.clear()
        for supplier in self.SupplierController.getAllSupplier():   
            self.supplier_list.addItem(supplier)
            
        for items in self.SupplierController.getAllItems():
            self.item_list.addItem(items)

        self.supplier_label.setText(self.supplier_str)
        self.supplier_location_label.setText(self.supplier_location_str)
        self.supply_label.setText(self.supply_str)
        self.transport_cost_label.setText(self.transport_cost_str)
        self.supply_cost_label.setText(self.supply_cost_str)
 
    def supplierListClicked(self):
        clicked_supplier = self.supplier_list.selectedItems()[0].text()
        self.setPath(clicked_supplier)

    def itemListClicked(self):
        pass

    def startCalculate(self):
        start_time = time.time()
        if(self.item_list.selectedItems() != []):
            lat , lon = self.customer_location.text().split(',')
            best_supplier_list = self.SupplierController.getBestSupplier(self.item_list.selectedItems()[0].text(), lat,lon)

            #อันดับที่ 1
            best_supplier = best_supplier_list[0][0]
            cost = best_supplier_list[0][1]
            
            #อันดับที่ 2
            best_supplier2 = best_supplier_list[1][0]
            cost2 = best_supplier_list[1][1]

            supplier_name2 = best_supplier2[0]
            supply_cost2 = best_supplier2[2]
            transport_cost2 = cost2
            distance2 = best_supplier2[1]
            
            #อันดับที่ 3
            best_supplier3 = best_supplier_list[2][0]
            cost3 = best_supplier_list[2][1]

            supplier_name3 = best_supplier3[0]
            supply_cost3 = best_supplier3[2]
            transport_cost3 = cost3
            distance3 = best_supplier3[1]

            #ไม่ยุ่ง
            best_supplier_cost = best_supplier[2]
            distance = best_supplier[1]
            best_supplier = best_supplier[0]

            self.setPath(best_supplier)
            
         ##setText
            self.supplier_label.setText(self.supplier_str + best_supplier)
            self.supply_label.setText(self.supply_str + self.SupplierController.getItem(best_supplier))
            self.transport_cost_label.setText(self.transport_cost_str + str("%.2f" % cost))
            self.supply_cost_label.setText(self.supply_cost_str + str(best_supplier_cost))
            self.distance_label.setText(self.distance_str + str(distance))

        ##setText2
            self.supplier_label_2.setText(self.supplier_str + supplier_name2)
            self.transport_cost_label_2.setText(self.transport_cost_str + str("%.2f" % supply_cost2))
            self.supply_cost_label_2.setText(self.supply_cost_str + str(transport_cost2))
            self.distance_label_2.setText(self.distance_str + str(distance2))

        ##setText3
            self.supplier_label_3.setText(self.supplier_str + supplier_name3)
            self.transport_cost_label_3.setText(self.transport_cost_str + str("%.2f" % supply_cost3))
            self.supply_cost_label_3.setText(self.supply_cost_str + str(transport_cost3))
            self.distance_label_3.setText(self.distance_str + str(distance3))
            
        Time = str(round((time.time() - start_time),2)) + "  วินาที"
        self.time_label.setText(self.time_str + Time)

    def setPath(self, supplier_name):
        location = self.SupplierController.getLocation(supplier_name)
        lat , lon = self.customer_location.text().split(',')
        self.MapController.setCustomerMarker(lat , lon)
        lat , lon = location.split(',')
        self.MapController.setSupplierMarker(lat,lon)
        self.webEngineView.setHtml(self.MapController.getHTML())
        
    def showUI(self):
        self.ui.show()
    

if __name__ == "__main__":
    app = QApplication(sys.argv)
    form = MainPage()
    form.showUI()
    sys.exit(app.exec_())

##AIzaSyBmt-IXSmfgH8AsEYAalEUgXuF23GCuNVQ
##resp = requests.get('https://maps.googleapis.com/maps/api/directions/json?origin=13.646667,100.681164&destination=13.730054,100.778890&key=AIzaSyBmt-IXSmfgH8AsEYAalEUgXuF23GCuNVQ')
##print(resp.json()["routes"][0]["legs"][0]["distance"]["text"])
